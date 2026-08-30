import csv
import io
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

FIELDS = [
    "Finding ID",
    "Target",
    "Asset",
    "Domain",
    "Subdomain",
    "Source Class",
    "Provider",
    "Query ID",
    "Query",
    "Result URL / Host",
    "Finding Category",
    "Finding Title",
    "Risk P1-P10",
    "Risk Score",
    "Confidence",
    "False Positive Probability",
    "Evidence IDs",
    "Technology",
    "Business Impact",
    "Technical Impact",
    "AI Provider",
    "AI Model",
    "AI Consensus",
    "Recommended Action",
    "Verification State",
    "First Seen",
    "Last Seen",
    "Owner",
    "Status",
]


def _row(item: dict):
    return [item.get(f, "") for f in FIELDS]


def export_json(findings: list[dict]) -> bytes:
    return json.dumps(
        {"schema_version": "2026.1", "findings": findings}, indent=2, default=str
    ).encode()


def export_csv(findings: list[dict]) -> bytes:
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(FIELDS)
    writer.writerows(_row(x) for x in findings)
    return out.getvalue().encode("utf-8-sig")


def export_xlsx(findings: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"
    ws.append(FIELDS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="163C42")
    for item in findings:
        ws.append(_row(item))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = min(
            45, max(12, max(len(str(c.value or "")) for c in column) + 2)
        )
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
